from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse, HttpResponseBadRequest
from django.utils import timezone

from booking.models import Booking
from special_tour.models import SpecialTourBooking
from transport.models import Bus, Fare

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def _parse_iso_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _resolve_period(period, start_date_raw, end_date_raw):
    today = timezone.localdate()

    if period == "7days":
        start_date = today - timedelta(days=6)
        return start_date, today, "Last 7 days"

    if period == "1month":
        start_date = today - timedelta(days=29)
        return start_date, today, "Last 1 month"

    if period == "overall":
        return None, None, "Overall"

    if period == "custom":
        start_date = _parse_iso_date(start_date_raw)
        end_date = _parse_iso_date(end_date_raw)
        if not start_date or not end_date:
            return None, None, None
        if start_date > end_date:
            return None, None, None
        return start_date, end_date, f"Custom ({start_date} to {end_date})"

    return None, None, None


def _calculate_booking_amount(booking):
    fare = (
        Fare.objects.filter(
            route=booking.bus.route,
            source_stop__name__iexact=booking.pickup,
            destination_stop__name__iexact=booking.dropoff,
        )
        .only("price")
        .first()
    )
    if fare is None:
        return Decimal("0.00")

    return fare.price * booking.passengers.count()


def _get_bookings_rows(queryset):
    rows = []
    total_amount = Decimal("0.00")

    for booking in queryset:
        passenger_count = booking.passengers.count()
        amount = _calculate_booking_amount(booking)
        total_amount += amount

        rows.append(
            {
                "id": booking.id,
                "created_at": timezone.localtime(booking.created_at).strftime("%Y-%m-%d %H:%M"),
                "travel_date": booking.travel_date.strftime("%Y-%m-%d"),
                "user": booking.user.username,
                "bus": booking.bus.number,
                "route": booking.bus.route.name,
                "pickup": booking.pickup,
                "dropoff": booking.dropoff,
                "passengers": passenger_count,
                "status": booking.status,
                "amount": amount,
            }
        )

    return rows, total_amount


def _get_special_tour_rows(queryset):
    rows = []
    total_estimated = Decimal("0.00")

    for tour in queryset:
        estimated = tour.estimated_price or Decimal("0.00")
        total_estimated += estimated

        rows.append(
            {
                "id": tour.id,
                "created_at": timezone.localtime(tour.created_at).strftime("%Y-%m-%d %H:%M"),
                "start_date": tour.journey_start_date.strftime("%Y-%m-%d"),
                "user": tour.user.username,
                "tour_type": tour.tour_type,
                "route": f"{tour.from_location} -> {tour.to_location}",
                "buses": tour.number_of_buses,
                "passengers": tour.passenger_count,
                "status": tour.status,
                "payment_status": tour.payment_status,
                "estimated": estimated,
            }
        )

    return rows, total_estimated


def _autosize_worksheet_columns(worksheet):
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 42)


def _write_summary_sheet(workbook, metadata, booking_rows, booking_total, special_rows, special_total):
    sheet = workbook.create_sheet("Summary")
    sheet.append(["NavaYatra Admin Report"])
    sheet.append([])
    sheet.append(["Generated At", metadata["generated_at"]])
    sheet.append(["Period", metadata["period_label"]])
    sheet.append(["Dataset", metadata["dataset_label"]])
    sheet.append(["Bus Filter", metadata["bus_label"]])
    sheet.append([])
    sheet.append(["Ticket Bookings Count", len(booking_rows)])
    sheet.append(["Ticket Bookings Revenue (Rs.)", float(booking_total)])
    sheet.append(["Special Tour Count", len(special_rows)])
    sheet.append(["Special Tour Estimated (Rs.)", float(special_total)])
    sheet.append(["Combined Value (Rs.)", float(booking_total + special_total)])

    for cell in sheet[1]:
        cell.font = Font(bold=True, size=14)


def _build_excel_response(metadata, booking_rows, booking_total, special_rows, special_total):
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_summary_sheet(
        workbook,
        metadata,
        booking_rows,
        booking_total,
        special_rows,
        special_total,
    )

    if booking_rows:
        booking_sheet = workbook.create_sheet("Ticket Bookings")
        booking_sheet.append(
            [
                "Booking ID",
                "Created At",
                "Travel Date",
                "User",
                "Bus",
                "Route",
                "Pickup",
                "Dropoff",
                "Passengers",
                "Status",
                "Amount (Rs.)",
            ]
        )

        for row in booking_rows:
            booking_sheet.append(
                [
                    row["id"],
                    row["created_at"],
                    row["travel_date"],
                    row["user"],
                    row["bus"],
                    row["route"],
                    row["pickup"],
                    row["dropoff"],
                    row["passengers"],
                    row["status"],
                    float(row["amount"]),
                ]
            )

        for cell in booking_sheet[1]:
            cell.font = Font(bold=True)

        _autosize_worksheet_columns(booking_sheet)

    if special_rows:
        special_sheet = workbook.create_sheet("Special Tours")
        special_sheet.append(
            [
                "Tour ID",
                "Created At",
                "Journey Start",
                "User",
                "Tour Type",
                "From-To",
                "No. Of Buses",
                "Passengers",
                "Status",
                "Payment",
                "Estimated (Rs.)",
            ]
        )

        for row in special_rows:
            special_sheet.append(
                [
                    row["id"],
                    row["created_at"],
                    row["start_date"],
                    row["user"],
                    row["tour_type"],
                    row["route"],
                    row["buses"],
                    row["passengers"],
                    row["status"],
                    row["payment_status"],
                    float(row["estimated"]),
                ]
            )

        for cell in special_sheet[1]:
            cell.font = Font(bold=True)

        _autosize_worksheet_columns(special_sheet)

    _autosize_worksheet_columns(workbook["Summary"])

    file_buffer = BytesIO()
    workbook.save(file_buffer)
    file_buffer.seek(0)

    filename = f"navayatra_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        file_buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_pdf_table(header, rows):
    table_data = [header] + rows
    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C62828")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CFCFCF")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF8F7")]),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("TOPPADDING", (0, 0), (-1, 0), 6),
            ]
        )
    )
    return table


def _build_pdf_response(metadata, booking_rows, booking_total, special_rows, special_total):
    pdf_buffer = BytesIO()
    document = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("NavaYatra Admin Report", styles["Title"]))
    story.append(Spacer(1, 8))
    story.append(
        Paragraph(
            (
                f"Generated At: {metadata['generated_at']}<br/>"
                f"Period: {metadata['period_label']}<br/>"
                f"Dataset: {metadata['dataset_label']}<br/>"
                f"Bus Filter: {metadata['bus_label']}"
            ),
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 10))

    summary_table = _build_pdf_table(
        ["Metric", "Value"],
        [
            ["Ticket Bookings Count", str(len(booking_rows))],
            ["Ticket Bookings Revenue (Rs.)", f"{booking_total:.2f}"],
            ["Special Tour Count", str(len(special_rows))],
            ["Special Tour Estimated (Rs.)", f"{special_total:.2f}"],
            ["Combined Value (Rs.)", f"{(booking_total + special_total):.2f}"],
        ],
    )
    story.append(summary_table)
    story.append(Spacer(1, 12))

    if booking_rows:
        story.append(Paragraph("Ticket Bookings", styles["Heading2"]))
        story.append(Spacer(1, 6))
        booking_table = _build_pdf_table(
            [
                "ID",
                "Created At",
                "Travel Date",
                "User",
                "Bus",
                "Route",
                "Pickup",
                "Dropoff",
                "Pax",
                "Status",
                "Amount",
            ],
            [
                [
                    row["id"],
                    row["created_at"],
                    row["travel_date"],
                    row["user"],
                    row["bus"],
                    row["route"],
                    row["pickup"],
                    row["dropoff"],
                    row["passengers"],
                    row["status"],
                    f"{row['amount']:.2f}",
                ]
                for row in booking_rows
            ],
        )
        story.append(booking_table)
        story.append(Spacer(1, 12))

    if special_rows:
        story.append(Paragraph("Special Tours", styles["Heading2"]))
        story.append(Spacer(1, 6))
        special_table = _build_pdf_table(
            [
                "ID",
                "Created At",
                "Start Date",
                "User",
                "Type",
                "From-To",
                "Buses",
                "Pax",
                "Status",
                "Payment",
                "Estimated",
            ],
            [
                [
                    row["id"],
                    row["created_at"],
                    row["start_date"],
                    row["user"],
                    row["tour_type"],
                    row["route"],
                    row["buses"],
                    row["passengers"],
                    row["status"],
                    row["payment_status"],
                    f"{row['estimated']:.2f}",
                ]
                for row in special_rows
            ],
        )
        story.append(special_table)

    document.build(story)
    pdf_buffer.seek(0)

    filename = f"navayatra_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response = HttpResponse(pdf_buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@staff_member_required
def admin_report_download(request):
    output_format = request.GET.get("format", "pdf").lower()
    dataset = request.GET.get("dataset", "combined").lower()
    period = request.GET.get("period", "7days").lower()
    bus_id_raw = request.GET.get("bus_id")
    start_date_raw = request.GET.get("start_date")
    end_date_raw = request.GET.get("end_date")

    if output_format not in {"pdf", "xlsx"}:
        return HttpResponseBadRequest("Invalid format")

    if dataset not in {"bookings", "special_tours", "combined"}:
        return HttpResponseBadRequest("Invalid dataset")

    start_date, end_date, period_label = _resolve_period(period, start_date_raw, end_date_raw)
    if period_label is None:
        return HttpResponseBadRequest("Invalid period or custom date range")

    bus = None
    if bus_id_raw:
        try:
            bus = Bus.objects.filter(id=int(bus_id_raw)).only("id", "number").first()
        except (TypeError, ValueError):
            return HttpResponseBadRequest("Invalid bus filter")

    booking_queryset = Booking.objects.select_related("user", "bus", "bus__route").prefetch_related("passengers")
    if start_date:
        booking_queryset = booking_queryset.filter(created_at__date__gte=start_date)
    if end_date:
        booking_queryset = booking_queryset.filter(created_at__date__lte=end_date)
    if bus is not None:
        booking_queryset = booking_queryset.filter(bus_id=bus.id)

    special_queryset = SpecialTourBooking.objects.select_related("user")
    if start_date:
        special_queryset = special_queryset.filter(created_at__date__gte=start_date)
    if end_date:
        special_queryset = special_queryset.filter(created_at__date__lte=end_date)

    booking_rows = []
    booking_total = Decimal("0.00")
    special_rows = []
    special_total = Decimal("0.00")

    if dataset in {"bookings", "combined"}:
        booking_rows, booking_total = _get_bookings_rows(booking_queryset.order_by("-created_at"))

    if dataset in {"special_tours", "combined"}:
        special_rows, special_total = _get_special_tour_rows(special_queryset.order_by("-created_at"))

    metadata = {
        "generated_at": timezone.localtime(timezone.now()).strftime("%Y-%m-%d %H:%M"),
        "period_label": period_label,
        "dataset_label": {
            "bookings": "Ticket Bookings",
            "special_tours": "Special Tours",
            "combined": "Combined",
        }[dataset],
        "bus_label": bus.number if bus is not None else "All Buses",
    }

    if output_format == "xlsx":
        return _build_excel_response(metadata, booking_rows, booking_total, special_rows, special_total)

    return _build_pdf_response(metadata, booking_rows, booking_total, special_rows, special_total)
